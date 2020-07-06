import openpyxl as xl

from openpyxl.chart import BarChart, Reference
import openpyxl.pivot as pivot

workbook = xl.load_workbook("transactions.xlsx")
sheet_name = workbook.sheetnames[0]
sheet = workbook[sheet_name]
for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    price = cell.value
    corrected_price = price * 0.9
    new_cell = sheet.cell(row, 4)
    new_cell.value = corrected_price

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)

sheet.add_chart(chart, 'e2')
workbook.save("Transactions2.xlsx")

