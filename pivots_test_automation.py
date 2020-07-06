import openpyxl as xl

import openpyxl.pivot as pivot

workbook = xl.load_workbook("transactions.xlsx")
sheet_name = workbook.sheetnames[0]
sheet = workbook[sheet_name]
for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    price = cell.value
    corrected_price = price * 0.9
    new_cell = sheet.cell(row, 4)
    new_cell.value = corrected_price

workbook.save("Transactions2.xlsx")

